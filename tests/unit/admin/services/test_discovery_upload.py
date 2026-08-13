import csv
import datetime
import io
from typing import Any

import openpyxl
import pytest

from statgpt.admin.services.discovery_upload import COLUMN_FIELDS, parse_discovery_file
from statgpt.admin.services.exceptions import DiscoveryUploadFormatError

_DATASETS_HEADERS = [
    "Reference area / country",
    "Regional coverage",
    "Excluded regional values",
    "Agency / organization",
    "Dataset ID",
    "Dataset name",
    "Description",
    "Dataset URL",
    "Time coverage",
    "Frequency coverage",
    "Indicators coverage (incl. units of measure)",
    "Relevant indicators not present in the dataset",
]

_ROW = [
    "Indonesia (IDN)",
    "None",
    "",
    "Bank Indonesia (BI)",
    "TABEL1_1",
    "Broad Money",
    "Money and banking table.",
    "https://www.bi.go.id/SEKI/tabel/TABEL1_1.xls",
    "From 1989-01 to 2026-06",
    "Monthly",
    "broad money (M2) (Rp billions)",
    "policy interest rates",
]


def _workbook(
    rows: list[list[Any]],
    headers: list[Any] | None = None,
    sheet_name: str = "Datasets",
    extra_sheet_first: bool = True,
) -> bytes:
    """Build a workbook in memory, so no binary fixture has to be committed."""
    wb = openpyxl.Workbook()
    first = wb.active
    assert first is not None
    if extra_sheet_first:
        # The template ships an `Instructions` sheet before `Datasets`.
        first.title = "Instructions"
        sheet = wb.create_sheet(sheet_name)
    else:
        first.title = sheet_name
        sheet = first

    sheet.append(_DATASETS_HEADERS if headers is None else headers)
    for row in rows:
        sheet.append(row)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _csv_bytes(rows: list[dict[str, Any]], fieldnames: list[str], delimiter: str = ",") -> bytes:
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=fieldnames, delimiter=delimiter)
    writer.writeheader()
    writer.writerows(rows)
    return buffer.getvalue().encode("utf-8")


def test_parses_template_headers_and_reports_cell_references() -> None:
    parsed = parse_discovery_file(_workbook([_ROW]), "book.xlsx", max_rows=100)

    assert len(parsed.rows) == 1
    row = parsed.rows[0]
    assert row.row_number == 2
    assert row.values["agency"] == "Bank Indonesia (BI)"
    assert row.values["missing_indicators"] == "policy interest rates"
    # Column D of the template, on the first data row.
    assert parsed.cell("agency", row.row_number) == "D2"


def test_accepts_the_instructions_sheet_spelling_of_column_l() -> None:
    """The template spells column L two ways; both have to resolve (see Excel notes)."""
    headers = _DATASETS_HEADERS[:-1] + ["Relevant indicators not present"]

    parsed = parse_discovery_file(_workbook([_ROW], headers=headers), "book.xlsx", max_rows=100)

    assert parsed.rows[0].values["missing_indicators"] == "policy interest rates"


def test_reads_second_sheet_when_datasets_sheet_is_absent() -> None:
    data = _workbook([_ROW], sheet_name="Sheet2")

    parsed = parse_discovery_file(data, "book.xlsx", max_rows=100)

    assert parsed.rows[0].values["dataset_id"] == "TABEL1_1"


def test_single_sheet_without_datasets_name_is_rejected() -> None:
    data = _workbook([_ROW], sheet_name="Sheet1", extra_sheet_first=False)

    with pytest.raises(DiscoveryUploadFormatError, match="no second sheet"):
        parse_discovery_file(data, "book.xlsx", max_rows=100)


def test_falls_back_to_column_order_when_headers_are_unrecognized() -> None:
    """A renamed header row is still readable, and is never imported as a record."""
    headers = [f"col{i}" for i in range(len(COLUMN_FIELDS))]

    parsed = parse_discovery_file(_workbook([_ROW], headers=headers), "book.xlsx", max_rows=100)

    assert len(parsed.rows) == 1
    assert parsed.rows[0].values["agency"] == "Bank Indonesia (BI)"


def test_partially_recognized_headers_missing_the_key_are_rejected() -> None:
    headers = ["Dataset name", "Description"] + [f"col{i}" for i in range(10)]

    with pytest.raises(DiscoveryUploadFormatError, match="missing required column"):
        parse_discovery_file(_workbook([_ROW], headers=headers), "book.xlsx", max_rows=100)


def test_two_columns_for_one_field_are_rejected() -> None:
    headers = _DATASETS_HEADERS + ["Agency / organization"]

    with pytest.raises(DiscoveryUploadFormatError, match="Two columns"):
        parse_discovery_file(_workbook([_ROW], headers=headers), "book.xlsx", max_rows=100)


def test_unknown_extra_columns_are_ignored() -> None:
    headers = _DATASETS_HEADERS + ["validation_status", "Some note"]

    parsed = parse_discovery_file(
        _workbook([_ROW + ["INVALID", "note"]], headers=headers), "book.xlsx", max_rows=100
    )

    assert parsed.rows[0].values["agency"] == "Bank Indonesia (BI)"
    assert "validation_status" not in parsed.rows[0].values


@pytest.mark.parametrize(
    "cell_value, expected",
    [
        (1960, "1960"),
        (1960.0, "1960"),
        (1.5, "1.5"),
        (True, "Yes"),
        (False, "No"),
        (None, ""),
        (datetime.datetime(2026, 5, 1), "2026-05-01"),
        (datetime.date(2026, 5, 1), "2026-05-01"),
    ],
)
def test_typed_cells_are_coerced_to_text(cell_value: object, expected: str) -> None:
    """openpyxl returns typed values, and Pydantic's lax mode will not coerce int -> str."""
    row: list = list(_ROW)
    row[8] = cell_value  # Time coverage

    parsed = parse_discovery_file(_workbook([row]), "book.xlsx", max_rows=100)

    assert parsed.rows[0].values["time_coverage"] == expected


def test_whitespace_is_normalized() -> None:
    """The agency cell below holds a literal U+00A0 after 'Bank', plus padding and runs."""
    row = list(_ROW)
    row[3] = "  Bank  Indonesia   (BI) "

    parsed = parse_discovery_file(_workbook([row]), "book.xlsx", max_rows=100)

    assert parsed.rows[0].values["agency"] == "Bank Indonesia (BI)"


def test_blank_rows_are_skipped_without_stopping_the_scan() -> None:
    """The template ships ~300 formatted-but-empty rows between and after the data."""
    blank: list = [None] * len(_DATASETS_HEADERS)
    whitespace_only = [" "] * len(_DATASETS_HEADERS)
    second = list(_ROW)
    second[4] = "TABEL1_2"

    parsed = parse_discovery_file(
        _workbook([_ROW, blank, whitespace_only, second, blank]), "book.xlsx", max_rows=100
    )

    assert [row.values["dataset_id"] for row in parsed.rows] == ["TABEL1_1", "TABEL1_2"]
    assert parsed.rows_skipped == 3
    # Row numbers still point at the real spreadsheet rows.
    assert [row.row_number for row in parsed.rows] == [2, 5]


def test_row_cap_is_enforced() -> None:
    rows = []
    for index in range(3):
        row = list(_ROW)
        row[4] = f"TABEL{index}"
        rows.append(row)

    with pytest.raises(DiscoveryUploadFormatError, match="more than 2 data rows"):
        parse_discovery_file(_workbook(rows), "book.xlsx", max_rows=2)


def test_legacy_xls_is_rejected_with_actionable_message() -> None:
    with pytest.raises(DiscoveryUploadFormatError, match="re-save it as .xlsx"):
        parse_discovery_file(b"\xd0\xcf\x11\xe0some legacy content", "book.xls", max_rows=100)


def test_corrupt_workbook_is_a_format_error_not_a_crash() -> None:
    with pytest.raises(DiscoveryUploadFormatError):
        parse_discovery_file(b"PK\x03\x04not really a zip", "book.xlsx", max_rows=100)


def test_unknown_file_type_is_rejected() -> None:
    with pytest.raises(DiscoveryUploadFormatError, match="Unsupported file type"):
        parse_discovery_file(b"%PDF-1.4", "book.pdf", max_rows=100)


def test_csv_with_field_names_round_trips_an_export() -> None:
    """A CSV exported by this application (snake_case headers) re-imports unchanged."""
    fieldnames = list(COLUMN_FIELDS)
    row = dict(zip(fieldnames, _ROW))

    parsed = parse_discovery_file(_csv_bytes([row], fieldnames), "discovery_datasets.csv", 100)

    assert parsed.rows[0].values == row
    # A CSV has no spreadsheet coordinates, so problems are reported by row only.
    assert parsed.cell("agency", 2) is None


def test_csv_with_workbook_headers_and_semicolon_delimiter() -> None:
    row = dict(zip(_DATASETS_HEADERS, _ROW))

    parsed = parse_discovery_file(
        _csv_bytes([row], _DATASETS_HEADERS, delimiter=";"), "book.csv", 100
    )

    assert parsed.rows[0].values["agency"] == "Bank Indonesia (BI)"
    assert parsed.rows[0].values["dataset_id"] == "TABEL1_1"


def test_csv_with_bom_and_cp1252_fallback() -> None:
    fieldnames = list(COLUMN_FIELDS)
    row = dict(zip(fieldnames, _ROW))
    row["agency"] = "Banco de España (BdE)"

    with_bom = "﻿".encode() + _csv_bytes([row], fieldnames)
    assert parse_discovery_file(with_bom, "x.csv", 100).rows[0].values["agency"] == row["agency"]

    as_cp1252 = _csv_bytes([row], fieldnames).decode().encode("cp1252")
    assert parse_discovery_file(as_cp1252, "x.csv", 100).rows[0].values["agency"] == row["agency"]


def test_empty_csv_is_rejected() -> None:
    with pytest.raises(DiscoveryUploadFormatError, match="empty"):
        parse_discovery_file(b"", "x.csv", 100)
