"""Build a Grade C discovery test workbook: the pilot records plus multi-country test rows.

The pilot workbook holds 81 records from one agency in one country, which cannot exercise the
country pre-filter at all. This adds rows that cover every filter case the MVP design names.

Every added record's dataset id is prefixed `TEST_`, so a fixture row that reaches a real
channel is identifiable in the index and in any referral the agent produces. URLs point at the
agency's statistics portal rather than at a specific table.
"""

import openpyxl

SRC = "Grade_C_Dataset_Discovery_Index_2026_08_10_1652.xlsx"
DST = "Grade_C_Discovery_Index_TEST_multi_country.xlsx"
FIRST_NEW_ROW = 83  # pilot data occupies rows 2-82

# (case, question, expectation) for the Test cases sheet, aligned with ROWS below.
ROWS: list[tuple[list[str], tuple[str, str, str]]] = [
    (
        [
            "Japan (JPN)", "None", "",
            "Bank of Japan (BOJ)", "TEST_BOJ_MB", "Monetary Base",
            "Monetary base and its components for Japan: banknotes in circulation, coins in "
            "circulation, and current account balances held at the Bank of Japan. Averages of "
            "daily figures and end-of-period figures are both published. Values in billions of yen.",
            "https://www.stat-search.boj.or.jp/",
            "From 1970-01 to present (latest available 2026-06)", "Monthly; Annual",
            "monetary base (in JPY billions); banknotes in circulation (in JPY billions); coins in "
            "circulation (in JPY billions); current account balances at the Bank of Japan (in JPY billions)",
            "consumer prices; trade in goods; gross domestic product, GDP",
        ],
        (
            "Single country, different from the pilot",
            "What is the monetary base of Japan?",
            "Retrieved and referred. Indonesian records must be filtered out.",
        ),
    ),
    (
        [
            "Japan (JPN)", "None", "",
            "Statistics Bureau of Japan (SBJ)", "TEST_SBJ_CPI", "Consumer Price Index",
            "Consumer price index for Japan, all items and by major group, including all items less "
            "fresh food. Published as an index and as year-on-year change. Values are revised once "
            "the annual rebasing is completed.",
            "https://www.stat.go.jp/english/data/cpi/",
            "From 1970-01 to present (latest available 2026-07)", "Monthly; Annual",
            "consumer prices (as an index (2020 = 100), as year-on-year % change); consumer prices "
            "excluding fresh food (as an index (2020 = 100), as year-on-year % change)",
            "producer prices; wages; monetary base",
        ],
        (
            "Two agencies in one country, different indicators",
            "What is inflation in Japan?",
            "The Statistics Bureau record is referred, not the Bank of Japan monetary records.",
        ),
    ),
    (
        [
            "Canada (CAN)", "provinces, territories", "Nunavut and Yukon are absent",
            "Statistics Canada (StatCan)", "TEST_STATCAN_LFS", "Labour Force Survey",
            "Labour force characteristics for Canada by province and territory: employment, "
            "unemployment, the unemployment rate, and the participation rate. Seasonally adjusted "
            "and unadjusted series are both published.",
            "https://www150.statcan.gc.ca/",
            "From 1976-01 to present (latest available 2026-07)", "Monthly; Annual",
            "unemployment rate (as %); employment (persons, thousands); unemployment (persons, "
            "thousands); labour force participation rate (as %)",
            "job vacancies; union membership; hours worked by industry",
        ],
        (
            "Excluded regional values populated (column C is empty in every pilot row)",
            "What is the unemployment rate in Nunavut?",
            "Referred with the exclusion stated, or ruled out. Not presented as covering Nunavut.",
        ),
    ),
    (
        [
            "Brazil (BRA)", "None", "",
            "Brazilian Institute of Geography and Statistics (IBGE)", "TEST_IBGE_IPCA",
            "Broad National Consumer Price Index (IPCA)",
            "Broad national consumer price index for Brazil, covering households in the main "
            "metropolitan areas, published as an index and as monthly and year-on-year change, "
            "with breakdowns by group of products and services.",
            "https://sidra.ibge.gov.br/",
            "From 1979-12 to present (latest available 2026-07)", "Monthly; Annual",
            "consumer prices (as an index (2020 = 100), as month-on-month % change, as "
            "year-on-year % change)",
            "producer prices; gross domestic product, GDP; labour force",
        ],
        (
            "Single country, no overlap with any other test record",
            "What is inflation in Brazil?",
            "Retrieved and referred. Nothing Indonesian or Japanese surfaces.",
        ),
    ),
    (
        [
            "India (IND)", "states, union territories", "",
            "Reserve Bank of India (RBI)", "TEST_RBI_MONEY", "Money and Banking Statistics",
            "Monetary aggregates and banking indicators for India: reserve money, broad money and "
            "its components, aggregate deposits, and bank credit to the commercial sector. Values "
            "in billions of rupees; recent periods may be provisional.",
            "https://dbie.rbi.org.in/",
            "From 1970-04 to present (latest available 2026-06)", "Monthly; Annual",
            "broad money M3 (in INR billions); reserve money (in INR billions); aggregate deposits "
            "(in INR billions); bank credit to the commercial sector (in INR billions)",
            "consumer prices; balance of payments; trade in goods",
        ],
        (
            "Single country sharing an indicator family with the pilot records",
            "What is broad money in India?",
            "Only the Indian record is referred. The 81 Indonesian money records are filtered out.",
        ),
    ),
    (
        [
            "Malaysia (MYS)", "None", "",
            "Bank Negara Malaysia (BNM)", "TEST_BNM_MONETARY", "Monetary and Financial Statistics",
            "Monetary and financial statistics for Malaysia: broad money and its components, "
            "banking system deposits, and loans outstanding by sector. Values in millions of ringgit.",
            "https://www.bnm.gov.my/statistics",
            "From 1990-01 to present (latest available 2026-06)", "Monthly",
            "broad money M3 (in MYR millions); narrow money M1 (in MYR millions); banking system "
            "deposits (in MYR millions); loans outstanding (in MYR millions)",
            "consumer prices; labour force; trade in services",
        ],
        (
            "Country adjacent to the pilot country, same indicator wording",
            "What is broad money in Malaysia?",
            "The Malaysian record wins. Retrieval alone would rank Indonesian records highly.",
        ),
    ),
    (
        [
            "Thailand (THA)", "None", "",
            "Bank of Thailand (BOT)", "TEST_BOT_MONETARY", "Monetary Statistics",
            "Monetary statistics for Thailand covering broad money, narrow money, and depository "
            "corporations' claims on the private sector. Values in millions of baht.",
            "https://www.bot.or.th/en/statistics.html",
            "From 1997-01 to present (latest available 2026-06)", "Monthly",
            "broad money (in THB millions); narrow money (in THB millions); claims on the private "
            "sector (in THB millions)",
            "consumer prices; gross domestic product, GDP",
        ],
        (
            "Single-country sibling of the multi-country record below",
            "What is broad money in Thailand?",
            "Both this record and the ASEAN multi-country record are legitimate candidates.",
        ),
    ),
    (
        [
            "Philippines (PHL)", "regions", "",
            "Philippine Statistics Authority (PSA)", "TEST_PSA_CPI", "Consumer Price Index",
            "Consumer price index for the Philippines, national and by region, published as an "
            "index and as year-on-year change, with breakdowns by commodity group.",
            "https://psa.gov.ph/statistics/price-indices",
            "From 1994-01 to present (latest available 2026-07)", "Monthly; Annual",
            "consumer prices (as an index (2018 = 100), as year-on-year % change); food price "
            "index (as an index (2018 = 100))",
            "producer prices; wages; monetary aggregates",
        ],
        (
            "Sub-national regional coverage without exclusions",
            "What is inflation in the Philippines by region?",
            "Retrieved and referred, with the regional breakdown visible in the referral.",
        ),
    ),
    (
        [
            "Germany (DEU)", "states", "",
            "Federal Statistical Office of Germany (Destatis)", "TEST_DESTATIS_NA",
            "National Accounts",
            "National accounts for Germany: gross domestic product by output, expenditure, and "
            "income approach, at current prices and chain-linked volumes, with seasonally adjusted "
            "quarterly series and state-level annual aggregates.",
            "https://www-genesis.destatis.de/",
            "From 1991 to present (latest available 2026-Q1)", "Quarterly; Annual",
            "gross domestic product, GDP (in current EUR billions, as chain-linked volume index "
            "(2020 = 100), as year-on-year % change); gross value added (in current EUR billions)",
            "consumer prices; labour force; monthly indicators",
        ],
        (
            "Euro-area member, tested against the group-label record below",
            "What is German GDP?",
            "This record is referred. The Euro area record may also surface through the sentinel.",
        ),
    ),
    (
        [
            "Indonesia (IDN); Malaysia (MYS); Thailand (THA); Philippines (PHL); Viet Nam (VNM); "
            "Singapore (SGP)",
            "None", "",
            "ASEAN Secretariat (ASEAN)", "TEST_ASEAN_MACRO", "Key Macroeconomic Indicators",
            "Comparable macroeconomic indicators for ASEAN member states: gross domestic product "
            "at current prices and in growth terms, consumer price inflation, population, and "
            "merchandise trade totals. Compiled from member states' national submissions.",
            "https://data.aseanstats.org/",
            "From 2005 to present (latest available 2025)", "Annual",
            "gross domestic product, GDP (in current USD billions, as year-on-year % change, per "
            "capita in current USD); consumer prices (as year-on-year % change); population "
            "(persons, millions); merchandise exports (in current USD billions); merchandise "
            "imports (in current USD billions)",
            "sub-national breakdowns; monthly indicators; trade in services",
        ],
        (
            "Multi-country cell — the case a whole-cell filter cannot match",
            "Compare GDP growth across ASEAN countries.",
            "Matched by a filter on any single member country. This row fails without array "
            "containment or one document per country.",
        ),
    ),
    (
        [
            "Euro area", "None", "",
            "Statistical Office of the European Union (Eurostat)", "TEST_ESTAT_HICP",
            "Harmonised Index of Consumer Prices",
            "Harmonised index of consumer prices for the euro area aggregate, published as an "
            "index and as annual rate of change, with breakdowns by the COICOP classification.",
            "https://ec.europa.eu/eurostat/data/database",
            "From 1996-01 to present (latest available 2026-07)", "Monthly; Annual",
            "consumer prices (as an index (2015 = 100), as year-on-year % change); consumer prices "
            "excluding energy and food (as an index (2015 = 100), as year-on-year % change)",
            "national accounts; labour force; sub-national data",
        ],
        (
            "Group label with no country code — the sentinel case",
            "What is inflation in Germany?",
            "Survives the country filter through the sentinel, then the judge decides whether a "
            "euro-area aggregate answers a question about Germany.",
        ),
    ),
    (
        [
            "World", "None", "",
            "International Monetary Fund (IMF)", "TEST_IMF_AGGREGATES",
            "Selected World and Regional Aggregates",
            "World and regional aggregates for output growth, consumer price inflation, and "
            "current account balances, including advanced economies and emerging market and "
            "developing economies groupings. Projections are published alongside outturns.",
            "https://data.imf.org/",
            "From 1980 to 2031 (projections from 2026)", "Annual",
            "gross domestic product, GDP (as year-on-year % change); consumer prices (as "
            "year-on-year % change); current account balance (as % of GDP)",
            "sub-national data; monthly indicators; country-level detail",
        ],
        (
            "World scope — the second sentinel case",
            "What is global inflation?",
            "Referred. Also survives any single-country filter, so it must not crowd out "
            "country-specific records.",
        ),
    ),
    (
        [
            "", "None", "",
            "Bank for International Settlements (BIS)", "TEST_BIS_CREDIT",
            "Credit to the Non-Financial Sector",
            "Credit to the non-financial sector across reporting economies: credit to the private "
            "non-financial sector, to households, and to non-financial corporations, expressed in "
            "national currency and as a percentage of gross domestic product.",
            "https://data.bis.org/",
            "From 1940-Q4 to present (latest available 2026-Q1)", "Quarterly",
            "credit to the private non-financial sector (as % of GDP); credit to households (as % "
            "of GDP); credit to non-financial corporations (as % of GDP)",
            "consumer prices; labour force; sub-national data",
        ],
        (
            "Empty reference area — the third sentinel case, and a blank-tolerance check",
            "What is household debt in Japan?",
            "Survives the filter through the sentinel. Confirms an empty cell is not treated as "
            "an unparseable failure that drops the record.",
        ),
    ),
    (
        [
            "Japan (JPN); Germany (DEU); partner countries: China; United States; European Union; "
            "United Kingdom",
            "None", "",
            "International Monetary Fund (IMF)", "TEST_IMF_TRADE_BILATERAL",
            "Bilateral Merchandise Trade, Selected Reporters",
            "Bilateral merchandise trade for selected reporting economies against their partner "
            "countries and partner groups, as reported and as derived estimates. Exports are free "
            "on board; imports are cost, insurance and freight.",
            "https://data.imf.org/",
            "From 1960-01 to present (latest available 2026-05)", "Monthly; Quarterly; Annual",
            "merchandise exports free on board (in current USD millions); merchandise imports "
            "cost, insurance and freight (in current USD millions); trade balance (in current USD "
            "millions)",
            "trade in services; trade by commodity; sub-national data",
        ],
        (
            "Bilateral record with the 'partner countries:' marker",
            "What does Japan export to China?",
            "Matched on the reporting country, Japan. Confirms parsing stops at the marker and "
            "partner names do not become filter values.",
        ),
    ),
    (
        [
            "Japan (JPN)", "None", "",
            "Bank of Japan (BOJ)", "TEST_BOJ_FX", "Foreign Exchange Rates",
            "Foreign exchange rates for the yen: spot rates against major currencies and the "
            "nominal and real effective exchange rate indices. Daily, monthly average, and "
            "end-of-period figures are published.",
            "https://www.stat-search.boj.or.jp/",
            "From 1980-01 to present (latest available 2026-07)", "Business daily; Monthly; Annual",
            "yen spot exchange rate (in JPY per USD, in JPY per EUR); nominal effective exchange "
            "rate (as an index (2020 = 100)); real effective exchange rate (as an index (2020 = 100))",
            "gross domestic product, GDP; consumer prices; interest rates; monetary base",
        ],
        (
            "Correct rejection — the right country, and column L names what will be asked",
            "What is Japan's GDP?",
            "Retrieval ranks this highly because column L contains the phrase. The judge must "
            "reject it and refer the Destatis-style national accounts record only if one exists "
            "for Japan, otherwise report nothing.",
        ),
    ),
    (
        [
            "Canada (CAN)", "provinces", "Quebec and Ontario are absent",
            "Statistics Canada (StatCan)", "TEST_STATCAN_RETAIL", "Retail Trade by Province",
            "Retail trade sales by province and by retail subsector, at current prices, "
            "seasonally adjusted and unadjusted. Recent months are subject to revision.",
            "https://www150.statcan.gc.ca/",
            "From 1991-01 to present (latest available 2026-06)", "Monthly; Annual",
            "retail sales (in current CAD millions); retail sales excluding motor vehicles and "
            "parts (in current CAD millions)",
            "e-commerce sales; wholesale trade; consumer prices",
        ],
        (
            "Correct rejection on region — right country and indicator, excluded region",
            "What were retail sales in Quebec?",
            "Ruled out for Quebec, or referred with the exclusion stated. Never presented as "
            "covering Quebec.",
        ),
    ),
]

HEADERS = (
    "Case",
    "Example question",
    "Expected outcome",
    "Agency / organization",
    "Dataset ID",
)


def main() -> None:
    workbook = openpyxl.load_workbook(SRC)
    sheet = workbook["Datasets"]

    cases = []
    for offset, entry in enumerate(ROWS):
        values, case = entry
        if len(values) != 12:
            raise ValueError(f"Row {offset} has {len(values)} columns, expected 12: {values[4:6]}")
        row = FIRST_NEW_ROW + offset
        for index, value in enumerate(values, start=1):
            sheet.cell(row, index).value = value
        cases.append((*case, values[3], values[4]))

    notes = workbook.create_sheet("Test cases")
    notes.append(list(HEADERS))
    for case in cases:
        notes.append(list(case))
    for column, width in zip("ABCDE", (52, 46, 78, 46, 28)):
        notes.column_dimensions[column].width = width

    workbook.save(DST)
    print(f"Wrote {DST}: {len(ROWS)} test records appended at rows "
          f"{FIRST_NEW_ROW}-{FIRST_NEW_ROW + len(ROWS) - 1}, plus a 'Test cases' sheet.")


if __name__ == "__main__":
    main()
