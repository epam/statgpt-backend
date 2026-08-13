"""Parsing of an uploaded discovery dataset file.

Pure and synchronous on purpose: this is the part with real branching, so it has to sit
where it can be unit-tested without a session or an HTTP app. Callers run it through
``asyncio.to_thread`` so a large workbook does not stall the event loop.

openpyxl is used directly rather than ``pandas.read_excel``: pandas coerces column dtypes
(``1960`` arrives as ``1960.0``), turns blanks into ``NaN``, mangles repeated headers, and
discards the cell coordinates the error report needs.
"""

import csv
import datetime
import io
import logging
import zipfile
from collections.abc import Iterable, Sequence
from dataclasses import dataclass, field
from enum import StrEnum, auto
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet._read_only import ReadOnlyWorksheet

from statgpt.common.utils import normalize_whitespace

from .exceptions import DiscoveryUploadFormatError

_log = logging.getLogger(__name__)

COLUMN_FIELDS: tuple[str, ...] = (
    "reference_area",  # A
    "regional_coverage",  # B
    "excluded_regional_values",  # C
    "agency",  # D
    "dataset_id",  # E
    "name",  # F
    "description",  # G
    "url",  # H
    "time_coverage",  # I
    "frequency_coverage",  # J
    "indicators_coverage",  # K
    "missing_indicators",  # L
)
"""The descriptive fields in workbook column order, used for the positional fallback."""

REQUIRED_FIELDS: tuple[str, ...] = ("agency", "dataset_id")
"""Halves of the natural key. A file without these columns cannot be reconciled at all."""

FIELD_LABELS: dict[str, str] = {
    "reference_area": "Reference area / country",
    "regional_coverage": "Regional coverage",
    "excluded_regional_values": "Excluded regional values",
    "agency": "Agency / organization",
    "dataset_id": "Dataset ID",
    "name": "Dataset name",
    "description": "Description",
    "url": "Dataset URL",
    "time_coverage": "Time coverage",
    "frequency_coverage": "Frequency coverage",
    "indicators_coverage": "Indicators coverage (incl. units of measure)",
    "missing_indicators": "Relevant indicators not present in the dataset",
}
"""Human labels, so an error message names the column the way the file does."""

_WORKBOOK_HEADERS: dict[str, str] = {
    "reference area / country": "reference_area",
    "regional coverage": "regional_coverage",
    "excluded regional values": "excluded_regional_values",
    "agency / organization": "agency",
    "dataset id": "dataset_id",
    "dataset name": "name",
    "description": "description",
    "dataset url": "url",
    "time coverage": "time_coverage",
    "frequency coverage": "frequency_coverage",
    "indicators coverage (incl. units of measure)": "indicators_coverage",
    "indicators coverage": "indicators_coverage",
    # The template spells column L two different ways: this one on the `Instructions`
    # sheet, the longer one on `Datasets`. Both have to resolve.
    "relevant indicators not present": "missing_indicators",
    "relevant indicators not present in the dataset": "missing_indicators",
}


def _build_header_aliases() -> dict[str, str]:
    """Header text (normalized + casefolded) -> field name.

    Carries the workbook headers plus the field names themselves, so a CSV exported by
    this application re-imports unchanged.
    """
    aliases = dict(_WORKBOOK_HEADERS)
    for name in COLUMN_FIELDS:
        aliases[name] = name
        aliases[name.replace("_", " ")] = name
        head, *rest = name.split("_")
        aliases[head + "".join(part.title() for part in rest)] = name
    return {key.casefold(): value for key, value in aliases.items()}


_HEADER_ALIASES: dict[str, str] = _build_header_aliases()

_DATASETS_SHEET_NAME: str = "datasets"
_XLSX_SIGNATURE = b"PK\x03\x04"
_XLS_SIGNATURE = b"\xd0\xcf\x11\xe0"


class _DiscoveryFileFormat(StrEnum):
    XLSX = auto()
    CSV = auto()


@dataclass(frozen=True)
class ParsedRow:
    """One data row of an uploaded file, already whitespace-normalized."""

    row_number: int
    """1-based row number in the file, so an error can point at it."""

    values: dict[str, str]
    """Field name -> value. Fields whose column is absent are not present here."""


@dataclass(frozen=True)
class ParsedFile:
    rows: list[ParsedRow] = field(default_factory=list)
    rows_skipped: int = 0
    """Blank rows skipped, such as the ~300 empty formatted rows the template ships."""

    column_letters: dict[str, str] = field(default_factory=dict)
    """Field name -> column letter, so a problem can be reported as a cell reference."""

    def cell(self, field_name: str, row_number: int) -> str | None:
        """The cell reference for a field on a row, e.g. 'D14'. None if not a spreadsheet."""
        letter = self.column_letters.get(field_name)
        return f"{letter}{row_number}" if letter else None


def _detect_format(data: bytes, filename: str | None = None) -> _DiscoveryFileFormat:
    """Classify an upload by its signature bytes, falling back to the filename extension."""
    if data[:4] == _XLSX_SIGNATURE:
        return _DiscoveryFileFormat.XLSX
    if data[:4] == _XLS_SIGNATURE:
        raise DiscoveryUploadFormatError(
            "The file is a legacy Excel workbook (.xls). Open it in Excel and re-save it"
            " as .xlsx, or export it as CSV."
        )

    name = (filename or "").casefold()
    if name.endswith(".csv") or name.endswith(".txt"):
        return _DiscoveryFileFormat.CSV
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        raise DiscoveryUploadFormatError(
            "The file is named like an Excel workbook but is not one. Re-save it as .xlsx."
        )
    raise DiscoveryUploadFormatError(
        "Unsupported file type. Upload the discovery workbook as .xlsx, or its data as .csv."
    )


def _coerce_cell(value: Any) -> str:
    """Render a cell as text.

    openpyxl returns typed values for typed cells, and Pydantic's lax mode does not
    coerce ``int -> str``, so the conversion has to happen here. Order matters: ``bool``
    is a subclass of ``int``, and ``1960.0`` has to render as ``1960``, not ``1960.0``.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        # These cells are read by a search layer and by a person, not by Python.
        return "Yes" if value else "No"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    if isinstance(value, datetime.datetime):
        return value.date().isoformat() if value.time() == datetime.time.min else value.isoformat()
    if isinstance(value, (datetime.date, datetime.time)):
        return value.isoformat()
    return str(value)


def _resolve_headers(header_cells: list[Any]) -> dict[int, str]:
    """Column index -> field name, from a header row.

    Unrecognized columns are ignored, so extra columns (including the status columns of an
    exported CSV) are tolerated. Two columns resolving to the same field is an error: one of
    them would be silently dropped.
    """
    resolved: dict[int, str] = {}
    seen: dict[str, int] = {}
    for index, cell in enumerate(header_cells):
        text = normalize_whitespace(_coerce_cell(cell))
        if not text:
            continue
        field_name = _HEADER_ALIASES.get(text.casefold())
        if field_name is None:
            continue
        if field_name in seen:
            raise DiscoveryUploadFormatError(
                f"Two columns describe {FIELD_LABELS[field_name]!r}"
                f" ({get_column_letter(seen[field_name] + 1)} and"
                f" {get_column_letter(index + 1)}). Remove one of them."
            )
        seen[field_name] = index
        resolved[index] = field_name
    return resolved


def _positional_headers(width: int) -> dict[int, str]:
    """Column index -> field name for a file whose header row could not be recognized.

    Only usable when the row is exactly as wide as the template. Row 1 is still treated as
    a header rather than as data: every real file has one, and importing a renamed header
    as a record would create a junk row.
    """
    if width != len(COLUMN_FIELDS):
        raise DiscoveryUploadFormatError(
            f"The header row does not name any known column. Expected the discovery"
            f" template's columns ({', '.join(FIELD_LABELS[f] for f in REQUIRED_FIELDS)},"
            f" ...) or exactly {len(COLUMN_FIELDS)} columns in template order,"
            f" found {width}."
        )
    _log.warning("Discovery upload header row unrecognized; falling back to column order A-L.")
    return dict(enumerate(COLUMN_FIELDS))


def _resolve_headers_or_positional(header_cells: list[Any]) -> dict[int, str]:
    resolved = _resolve_headers(header_cells)
    missing = [name for name in REQUIRED_FIELDS if name not in resolved.values()]
    if not missing:
        return resolved
    if resolved:
        raise DiscoveryUploadFormatError(
            "The file is missing required column(s): "
            + ", ".join(repr(FIELD_LABELS[name]) for name in missing)
            + "."
        )
    return _positional_headers(len(header_cells))


def _build_rows(
    data_rows: Iterable[tuple[int, Sequence[Any]]],
    headers: dict[int, str],
    max_rows: int,
) -> tuple[list[ParsedRow], int]:
    """Normalize the data rows of a file, stopping as soon as the row cap is exceeded.

    Takes an iterable rather than a list so the cap bounds memory: a sheet is streamed and
    refused while it is being read, instead of after it has been materialized in full. The
    byte cap on the upload is no substitute - xlsx is compressed XML, so a small file can
    expand by orders of magnitude.
    """
    rows: list[ParsedRow] = []
    skipped = 0

    for row_number, cells in data_rows:
        values = {
            field_name: normalize_whitespace(_coerce_cell(cells[index]))
            for index, field_name in headers.items()
            if index < len(cells)
        }
        if not any(values.values()):
            # The template ships hundreds of formatted-but-empty rows, and `ws.max_row`
            # is inflated in read-only mode, so blanks are skipped rather than stopped at.
            skipped += 1
            continue
        rows.append(ParsedRow(row_number=row_number, values=values))
        if len(rows) > max_rows:
            raise DiscoveryUploadFormatError(
                f"The file has more than {max_rows} data rows. Split it and upload the parts."
            )

    return rows, skipped


def _column_letters(headers: dict[int, str]) -> dict[str, str]:
    return {name: get_column_letter(index + 1) for index, name in headers.items()}


def _get_worksheet(workbook: Workbook) -> ReadOnlyWorksheet:
    """Pick the sheet holding the records."""
    worksheet = None
    for sheet_name in workbook.sheetnames:
        if sheet_name.strip().casefold() == _DATASETS_SHEET_NAME:
            worksheet = workbook[sheet_name]
            break
    if worksheet is None:
        if len(workbook.sheetnames) < 2:
            raise DiscoveryUploadFormatError(
                f"The workbook has no {_DATASETS_SHEET_NAME.title()!r} sheet and no second"
                f" sheet to read instead. Upload the filled discovery template."
            )
        worksheet = workbook[workbook.sheetnames[1]]
    return worksheet


def _parse_workbook(data: bytes, max_rows: int) -> ParsedFile:
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except (zipfile.BadZipFile, InvalidFileException, KeyError, ValueError) as e:
        raise DiscoveryUploadFormatError(
            "The workbook could not be read. It may be corrupted or password-protected;"
            " re-save it as .xlsx and try again."
        ) from e

    try:
        worksheet = _get_worksheet(workbook)

        rows_iter = worksheet.values
        try:
            header_cells = list(next(rows_iter))
        except StopIteration as e:
            raise DiscoveryUploadFormatError("The sheet is empty.") from e

        headers = _resolve_headers_or_positional(header_cells)
        # Consumed here, inside the `try`: streaming is what makes `max_rows` bound memory,
        # and the rows cannot be read once the `finally` below has closed the workbook.
        rows, skipped = _build_rows(enumerate(rows_iter, start=2), headers, max_rows)
    finally:
        workbook.close()

    return ParsedFile(rows=rows, rows_skipped=skipped, column_letters=_column_letters(headers))


def _decode_csv(data: bytes) -> str:
    for encoding in ("utf-8-sig", "cp1252"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise DiscoveryUploadFormatError(
        "The CSV file's encoding could not be determined. Re-save it as UTF-8."
    )


def _csv_dialect(text: str) -> type[csv.Dialect] | csv.Dialect:
    """Sniff the delimiter, since Excel writes ';' in some locales."""
    sample = text[:8192]
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        return csv.excel


def _parse_csv(data: bytes, max_rows: int) -> ParsedFile:
    text = _decode_csv(data)
    reader = csv.reader(io.StringIO(text, newline=""), dialect=_csv_dialect(text))

    try:
        header_cells = next(reader)
    except StopIteration as e:
        raise DiscoveryUploadFormatError("The CSV file is empty.") from e

    headers = _resolve_headers_or_positional(header_cells)
    rows, skipped = _build_rows(enumerate(reader, start=2), headers, max_rows)
    # A CSV has no spreadsheet coordinates, so problems are reported by row only.
    return ParsedFile(rows=rows, rows_skipped=skipped)


def parse_discovery_file(data: bytes, filename: str | None, max_rows: int) -> ParsedFile:
    """Parse an uploaded workbook or CSV into normalized rows.

    Raises `DiscoveryUploadFormatError` for anything that is not a readable discovery
    file - the caller maps that to a 400.
    """
    file_format = _detect_format(data, filename)
    _log.info(f"Parsing discovery upload {filename!r} as {file_format}.")

    if file_format is _DiscoveryFileFormat.XLSX:
        return _parse_workbook(data, max_rows)
    return _parse_csv(data, max_rows)
